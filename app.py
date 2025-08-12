from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, make_response
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
import plotly.graph_objs as go
import plotly.utils
import json
from collections import defaultdict
from functools import wraps
import pymysql
import os
from dotenv import load_dotenv
from werkzeug.utils import secure_filename
import uuid
import csv
import io
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.pdfgen import canvas
    from reportlab.graphics.shapes import Drawing, Rect, String
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.graphics.widgetbase import Widget
    from reportlab.graphics import renderPDF
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

# 加载环境变量
load_dotenv()

pymysql.install_as_MySQLdb()

app = Flask(__name__)

# 从环境变量加载配置
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'your-secret-key-here')
app.config['SQLALCHEMY_DATABASE_URI'] = f"mysql://{os.getenv('DB_USER', '')}:{os.getenv('DB_PASSWORD', '')}@{os.getenv('DB_HOST', '')}/{os.getenv('DB_NAME', '')}?charset=utf8mb4&connect_timeout=60&read_timeout=60&write_timeout=60"
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_size': 10,
    'pool_timeout': 20,
    'pool_recycle': 3600,
    'max_overflow': 20,
    'pool_pre_ping': True
}

# 文件上传配置
app.config['UPLOAD_FOLDER'] = 'static/uploads/suppliers'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

# 确保上传目录存在
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

@login_manager.unauthorized_handler
def unauthorized():
    # 检查是否是 AJAX 请求
    # if request.headers.get('Accept') == 'application/json':
    #     # return jsonify({'error': '请先登录才能访问此页面'}), 401
    # # flash('请先登录才能访问此页面', 'error')
    return redirect(url_for('login'))

# 数据模型
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(120), nullable=False)
    role = db.Column(db.String(20), nullable=False, default='employee')
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Category(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), nullable=False)
    type = db.Column(db.String(10), nullable=False)
    color = db.Column(db.String(7), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)

class Supplier(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    contact_person = db.Column(db.String(50))
    phone = db.Column(db.String(20))
    address = db.Column(db.String(200))
    supplier_type = db.Column(db.String(50))  # 供应商类型：生鲜供应商、包装材料供应商、设备供应商、服务供应商
    products = db.Column(db.Text)  # 产品信息
    image_path = db.Column(db.String(255))  # 保留原字段用于兼容
    supply_method = db.Column(db.String(50))  # 供应方式：产地直供/合作社、本地批发市场供应商、一件代发供应商、社区本地农户/小农场
    notes = db.Column(db.Text)  # 备注信息
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    is_active = db.Column(db.Boolean, default=True)

class SupplierImage(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    supplier_id = db.Column(db.Integer, db.ForeignKey('supplier.id'), nullable=False)
    image_path = db.Column(db.String(255), nullable=False)
    upload_time = db.Column(db.DateTime, default=datetime.utcnow)
    supplier = db.relationship('Supplier', backref='images')



class Product(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    category = db.Column(db.String(50))
    supplier_id = db.Column(db.Integer, db.ForeignKey('supplier.id'))
    cost_price = db.Column(db.Float)
    selling_price = db.Column(db.Float)
    unit = db.Column(db.String(20))
    stock = db.Column(db.Integer, default=0)  # 库存数量
    description = db.Column(db.Text)  # 商品描述
    image_path = db.Column(db.String(255))  # 商品图片路径
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    is_active = db.Column(db.Boolean, default=True)
    supplier = db.relationship('Supplier', backref='product_list')

class Receivable(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    receivable_number = db.Column(db.String(50), unique=True, nullable=False)  # 应收款单号
    title = db.Column(db.String(200), nullable=False)  # 应收款标题/客户名称
    amount = db.Column(db.Float, nullable=False)  # 应收金额
    received_amount = db.Column(db.Float, default=0.0)  # 已收金额
    status = db.Column(db.String(20), default='pending')  # pending, partial, received
    invoice_date = db.Column(db.Date)  # 开票日期
    due_date = db.Column(db.Date)  # 到期日期
    payment_terms = db.Column(db.Integer, default=30)  # 账期天数
    contact_person = db.Column(db.String(100))  # 联系人
    contact_phone = db.Column(db.String(20))  # 联系电话
    contact_address = db.Column(db.String(500))  # 联系地址
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    received_at = db.Column(db.DateTime)
    notes = db.Column(db.Text)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    user = db.relationship('User', backref='receivables')
    
    @property
    def remaining_amount(self):
        """计算剩余未收金额"""
        return self.amount - self.received_amount
    
    @property
    def overdue_days(self):
        """计算逾期天数"""
        if not self.due_date or self.status == 'received':
            return 0
        today = datetime.now().date()
        if today > self.due_date:
            return (today - self.due_date).days
        return 0
    
    @property
    def days_until_due(self):
        """计算距离到期的天数"""
        if not self.due_date or self.status == 'received':
            return None
        today = datetime.now().date()
        if today <= self.due_date:
            return (self.due_date - today).days
        return 0

class Transaction(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    amount = db.Column(db.Float, nullable=False)
    type = db.Column(db.String(10), nullable=False)
    description = db.Column(db.String(200))
    supplier_description = db.Column(db.String(200))  # 供应商/描述字段
    date = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    category_id = db.Column(db.Integer, db.ForeignKey('category.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    supplier_id = db.Column(db.Integer, db.ForeignKey('supplier.id'))
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'))
    quantity = db.Column(db.Float)
    unit_price = db.Column(db.Float)

    category = db.relationship('Category', backref='transactions')
    supplier = db.relationship('Supplier', backref='transactions')
    product = db.relationship('Product', backref='transactions')

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

def admin_required(f):
    @wraps(f)
    @login_required
    def decorated_function(*args, **kwargs):
        if not current_user.role in ['admin', 'manager']:
            flash('权限不足', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def edit_required(f):
    @wraps(f)
    @login_required
    def decorated_function(*args, **kwargs):
        if current_user.role == 'employee':
            flash('员工权限不足，只能查看数据', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/')
@app.route('/dashboard')
@login_required
def dashboard():
    today = datetime.now().date()
    month_start = today.replace(day=1)
    
    # 本月收支统计
    month_income = db.session.query(db.func.sum(Transaction.amount)).filter(
        Transaction.type == 'income',
        Transaction.date >= month_start
    ).scalar() or 0
    
    month_expense = db.session.query(db.func.sum(Transaction.amount)).filter(
        Transaction.type == 'expense',
        Transaction.date >= month_start
    ).scalar() or 0
    
    # 总体统计数据
    total_transactions = Transaction.query.count()
    total_users = User.query.count()
    total_suppliers = Supplier.query.count()
    total_products = Product.query.count()
    total_categories = Category.query.count()
    
    # 活跃数据统计
    active_suppliers = Supplier.query.count()
    active_products = Product.query.filter_by(is_active=True).count()
    active_users = User.query.filter_by(is_active=True).count()
    
    # 总收入支出统计
    total_income = db.session.query(db.func.sum(Transaction.amount)).filter(
        Transaction.type == 'income'
    ).scalar() or 0
    
    total_expense = db.session.query(db.func.sum(Transaction.amount)).filter(
        Transaction.type == 'expense'
    ).scalar() or 0
    
    # 最近交易
    recent_transactions = Transaction.query.order_by(Transaction.date.desc()).limit(10).all()
    
    # 分类统计
    category_stats = db.session.query(
        Category.name,
        db.func.sum(Transaction.amount).label('total')
    ).join(Transaction).filter(
        Transaction.type == 'expense',
        Transaction.date >= month_start
    ).group_by(Category.name).all()
    
    # 供应商统计
    supplier_stats = db.session.query(
        Supplier.name,
        db.func.sum(Transaction.amount).label('total')
    ).join(Transaction).filter(
        Transaction.type == 'expense',
        Transaction.date >= month_start
    ).group_by(Supplier.name).limit(5).all()
    
    # 应收款统计
    receivables_list = Receivable.query.filter_by(user_id=current_user.id).all()
    total_receivables = sum(r.amount for r in receivables_list)
    received_receivables = sum(r.amount for r in receivables_list if r.status == 'received')
    pending_receivables = total_receivables - received_receivables
    
    # 计算现金余额（总收入 - 总支出）
    cash_balance = total_income - total_expense
    
    # 获取过去12个月的收支数据用于趋势图
    trend_data = []
    for i in range(12):
        month_date = (today.replace(day=1) - timedelta(days=32*i)).replace(day=1)
        next_month = (month_date.replace(day=28) + timedelta(days=4)).replace(day=1)
        
        month_income_trend = db.session.query(db.func.sum(Transaction.amount)).filter(
            Transaction.type == 'income',
            Transaction.date >= month_date,
            Transaction.date < next_month
        ).scalar() or 0
        
        month_expense_trend = db.session.query(db.func.sum(Transaction.amount)).filter(
            Transaction.type == 'expense',
            Transaction.date >= month_date,
            Transaction.date < next_month
        ).scalar() or 0
        
        trend_data.insert(0, {
            'month': month_date.strftime('%m月'),
            'income': float(month_income_trend),
            'expense': float(month_expense_trend)
        })
    
    # 获取收入分类数据用于饼图
    income_category_stats = db.session.query(
        Category.name,
        db.func.sum(Transaction.amount).label('total')
    ).join(Transaction).filter(
        Transaction.type == 'income',
        Transaction.date >= month_start
    ).group_by(Category.name).all()
    
    # 获取支出分类数据用于饼图
    expense_category_stats = db.session.query(
        Category.name,
        db.func.sum(Transaction.amount).label('total')
    ).join(Transaction).filter(
        Transaction.type == 'expense',
        Transaction.date >= month_start
    ).group_by(Category.name).all()
    
    # 本年收支统计
    year_start = today.replace(month=1, day=1)
    year_income = db.session.query(db.func.sum(Transaction.amount)).filter(
        Transaction.type == 'income',
        Transaction.date >= year_start
    ).scalar() or 0
    
    year_expense = db.session.query(db.func.sum(Transaction.amount)).filter(
        Transaction.type == 'expense',
        Transaction.date >= year_start
    ).scalar() or 0
    
    # 获取年度收入分类数据用于饼图
    year_income_category_stats = db.session.query(
        Category.name,
        db.func.sum(Transaction.amount).label('total')
    ).join(Transaction).filter(
        Transaction.type == 'income',
        Transaction.date >= year_start
    ).group_by(Category.name).all()
    
    # 获取年度支出分类数据用于饼图
    year_expense_category_stats = db.session.query(
        Category.name,
        db.func.sum(Transaction.amount).label('total')
    ).join(Transaction).filter(
        Transaction.type == 'expense',
        Transaction.date >= year_start
    ).group_by(Category.name).all()
    
    # 获取年度数据用于趋势图
    current_year = today.year
    year_data = []
    
    # 获取过去6年的数据
    for year_offset in range(5, -1, -1):
        target_year = current_year - year_offset
        year_start_trend = datetime(target_year, 1, 1)
        year_end_trend = datetime(target_year + 1, 1, 1)
        
        year_income_trend = db.session.query(db.func.sum(Transaction.amount)).filter(
            Transaction.type == 'income',
            Transaction.date >= year_start_trend,
            Transaction.date < year_end_trend
        ).scalar() or 0
        
        year_expense_trend = db.session.query(db.func.sum(Transaction.amount)).filter(
            Transaction.type == 'expense',
            Transaction.date >= year_start_trend,
            Transaction.date < year_end_trend
        ).scalar() or 0
        
        year_data.append({
            'year': str(target_year),
            'income': float(year_income_trend),
            'expense': float(year_expense_trend)
        })
    
    return render_template('dashboard.html', 
                         month_income=month_income,
                         month_expense=month_expense,
                         year_income=year_income,
                         year_expense=year_expense,
                         recent_transactions=recent_transactions,
                         category_stats=category_stats,
                         supplier_stats=supplier_stats,
                         total_transactions=total_transactions,
                         total_users=total_users,
                         total_suppliers=total_suppliers,
                         total_products=total_products,
                         total_categories=total_categories,
                         active_suppliers=active_suppliers,
                         active_products=active_products,
                         active_users=active_users,
                         total_income=total_income,
                         total_expense=total_expense,
                         total_receivables=total_receivables,
                         received_receivables=received_receivables,
                         pending_receivables=pending_receivables,
                         cash_balance=cash_balance,
                         trend_data=trend_data,
                         year_data=year_data,
                         income_category_stats=income_category_stats,
                         expense_category_stats=expense_category_stats,
                         year_income_category_stats=year_income_category_stats,
                         year_expense_category_stats=year_expense_category_stats)

@app.route('/transactions')
@login_required
def transactions():
    page = request.args.get('page', 1, type=int)
    
    # 获取筛选参数
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    transaction_type = request.args.get('type')
    category_id = request.args.get('category_id')
    
    # 构建查询
    query = Transaction.query
    
    # 日期范围筛选
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
            query = query.filter(Transaction.date >= start_date_obj)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
            # 包含结束日期的整天
            end_date_obj = end_date_obj.replace(hour=23, minute=59, second=59)
            query = query.filter(Transaction.date <= end_date_obj)
        except ValueError:
            pass
    
    # 交易类型筛选
    if transaction_type and transaction_type in ['income', 'expense']:
        query = query.filter(Transaction.type == transaction_type)
    
    # 分类筛选
    if category_id:
        try:
            category_id_int = int(category_id)
            query = query.filter(Transaction.category_id == category_id_int)
        except ValueError:
            pass
    
    # 排序和分页
    transactions = query.order_by(Transaction.date.desc()).paginate(
        page=page, per_page=20, error_out=False)
    
    categories = Category.query.all()
    suppliers = Supplier.query.all()
    products = Product.query.filter_by(is_active=True).all()
    
    # 传递当前筛选参数到模板
    filter_params = {
        'start_date': start_date,
        'end_date': end_date,
        'type': transaction_type,
        'category_id': category_id
    }
    
    return render_template('transactions.html', 
                         transactions=transactions, 
                         categories=categories, 
                         suppliers=suppliers, 
                         products=products,
                         filter_params=filter_params)

@app.route('/export_transactions')
@login_required
def export_transactions():
    """导出收支记录为CSV文件"""
    # 获取筛选参数
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    transaction_type = request.args.get('type')
    category_id = request.args.get('category_id')
    export_format = request.args.get('format', 'csv')  # 默认CSV格式
    
    # 构建查询（与transactions路由相同的逻辑）
    query = Transaction.query
    
    # 日期范围筛选
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
            query = query.filter(Transaction.date >= start_date_obj)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
            end_date_obj = end_date_obj.replace(hour=23, minute=59, second=59)
            query = query.filter(Transaction.date <= end_date_obj)
        except ValueError:
            pass
    
    # 交易类型筛选
    if transaction_type and transaction_type in ['income', 'expense']:
        query = query.filter(Transaction.type == transaction_type)
    
    # 分类筛选
    if category_id:
        try:
            category_id_int = int(category_id)
            query = query.filter(Transaction.category_id == category_id_int)
        except ValueError:
            pass
    
    # 获取所有符合条件的交易记录
    transactions = query.order_by(Transaction.date.desc()).all()
    
    # 检查是否有数据
    if not transactions:
        flash('没有符合条件的交易记录可以导出', 'warning')
        return redirect(url_for('transactions'))
    
    if export_format == 'csv':
        # 创建CSV文件
        output = io.StringIO()
        writer = csv.writer(output)
        
        # 写入表头
        writer.writerow(['交易日期', '交易类型', '分类', '金额', '备注信息'])
        
        # 写入数据
        for transaction in transactions:
            writer.writerow([
                transaction.date.strftime('%Y-%m-%d'),
                '收入' if transaction.type == 'income' else '支出',
                transaction.category.name,
                transaction.amount,
                transaction.description or ''
            ])
        
        # 创建响应
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        
        # 生成文件名
        filename = f'transactions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        filename_utf8 = f'收支记录_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{filename_utf8.encode("utf-8").decode("latin1")}"'
        
        return response
    
    elif export_format == 'excel':
        if not EXCEL_AVAILABLE:
            flash('Excel导出功能不可用，请安装openpyxl库', 'error')
            return redirect(url_for('transactions'))
        
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "收支记录"
        
        # 设置表头
        headers = ['交易日期', '交易类型', '分类', '金额', '备注信息']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入数据
        for row, transaction in enumerate(transactions, 2):
            ws.cell(row=row, column=1, value=transaction.date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=2, value='收入' if transaction.type == 'income' else '支出')
            ws.cell(row=row, column=3, value=transaction.category.name)
            ws.cell(row=row, column=4, value=transaction.amount)
            ws.cell(row=row, column=5, value=transaction.description or '')
            
            # 设置交易类型颜色
            type_cell = ws.cell(row=row, column=2)
            if transaction.type == 'income':
                type_cell.font = Font(color="00B050")
            else:
                type_cell.font = Font(color="FF0000")
        
        # 调整列宽
        column_widths = [12, 10, 15, 12, 35]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 创建响应
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        # 生成文件名
        filename = f'transactions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        filename_utf8 = f'收支记录_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{filename_utf8.encode("utf-8").decode("latin1")}"'
        
        return response
    
    else:
        flash('不支持的导出格式', 'error')
        return redirect(url_for('transactions'))

@app.route('/add_transaction', methods=['GET', 'POST'])
@edit_required
def add_transaction():
    if request.method == 'POST':
        amount = float(request.form['amount'])
        transaction_type = request.form['type']
        description = request.form['description']
        supplier_description = request.form.get('supplier_description', '')
        category_id = int(request.form['category_id'])
        date = datetime.strptime(request.form['date'], '%Y-%m-%d')
        
        supplier_id = request.form.get('supplier_id') or None
        product_id = request.form.get('product_id') or None
        quantity = request.form.get('quantity') or None
        unit_price = request.form.get('unit_price') or None
        
        if supplier_id:
            supplier_id = int(supplier_id)
        if product_id:
            product_id = int(product_id)
        if quantity:
            quantity = float(quantity)
        if unit_price:
            unit_price = float(unit_price)
        
        transaction = Transaction(
            amount=amount,
            type=transaction_type,
            description=description,
            supplier_description=supplier_description,
            category_id=category_id,
            user_id=current_user.id,
            date=date,
            supplier_id=supplier_id,
            product_id=product_id,
            quantity=quantity,
            unit_price=unit_price,

        )
        db.session.add(transaction)
        db.session.commit()
        flash('交易记录添加成功！', 'success')
        return redirect(url_for('transactions'))
    
    # 获取所有分类，让前端JavaScript根据类型过滤
    transaction_type = request.args.get('type', 'expense')
    categories = Category.query.all()
    
    suppliers = Supplier.query.all()
    products = Product.query.filter_by(is_active=True).all()
    today = datetime.now().strftime('%Y-%m-%d')
    return render_template('add_transaction.html', categories=categories, suppliers=suppliers, products=products, today=today, transaction_type=transaction_type)

@app.route('/edit_transaction/<int:id>', methods=['GET', 'POST'])
@edit_required
def edit_transaction(id):
    transaction = Transaction.query.get_or_404(id)
    
    if request.method == 'POST':
        transaction.amount = float(request.form['amount'])
        transaction.type = request.form['type']
        transaction.description = request.form['description']
        transaction.supplier_description = request.form.get('supplier_description', '')
        transaction.category_id = int(request.form['category_id'])
        transaction.date = datetime.strptime(request.form['date'], '%Y-%m-%d')
        


        
        db.session.commit()
        flash('交易记录更新成功！', 'success')
        return redirect(url_for('transactions'))
    
    categories = Category.query.all()
    suppliers = Supplier.query.all()
    products = Product.query.filter_by(is_active=True).all()
    return render_template('edit_transaction.html', transaction=transaction, categories=categories, suppliers=suppliers, products=products)

@app.route('/delete_transaction/<int:id>', methods=['GET', 'POST'])
@edit_required
def delete_transaction(id):
    transaction = Transaction.query.get_or_404(id)
    
    db.session.delete(transaction)
    db.session.commit()
    flash('交易记录删除成功！', 'success')
    
    return redirect(url_for('transactions'))

@app.route('/categories')
@login_required
def categories():
    categories = Category.query.all()
    return render_template('categories.html', categories=categories)

@app.route('/add_category', methods=['GET', 'POST'])
@edit_required
def add_category():
    if request.method == 'POST':
        name = request.form['name']
        category_type = request.form['type']
        color = request.form['color']
        
        category = Category(
            name=name,
            type=category_type,
            color=color,
            user_id=current_user.id
        )
        db.session.add(category)
        db.session.commit()
        flash('分类添加成功！', 'success')
        return redirect(url_for('categories'))
    
    return render_template('add_category.html')

@app.route('/edit_category/<int:id>', methods=['GET', 'POST'])
@edit_required
def edit_category(id):
    category = Category.query.get_or_404(id)
    
    if request.method == 'POST':
        category.name = request.form['name']
        category.type = request.form['type']
        category.color = request.form['color']
        
        db.session.commit()
        flash('分类更新成功！', 'success')
        return redirect(url_for('categories'))
    
    return render_template('edit_category.html', category=category)

@app.route('/delete_category/<int:id>')
@edit_required
def delete_category(id):
    category = Category.query.get_or_404(id)
    
    # 检查是否有关联的交易
    if category.transactions:
        flash('无法删除：该分类下有关联的交易记录', 'error')
    else:
        db.session.delete(category)
        db.session.commit()
        flash('分类删除成功！', 'success')
    
    return redirect(url_for('categories'))

@app.route('/suppliers')
@login_required
def suppliers():
    suppliers = Supplier.query.all()
    return render_template('suppliers.html', suppliers=suppliers)

@app.route('/supplier/<int:id>')
@login_required
def supplier_detail(id):
    supplier = Supplier.query.get_or_404(id)
    return render_template('supplier_detail.html', supplier=supplier)

@app.route('/add_supplier', methods=['GET', 'POST'])
@edit_required
def add_supplier():
    if request.method == 'POST':
        name = request.form['name']
        contact_person = request.form['contact_person']
        phone = request.form['phone']
        address = request.form['address']
        supplier_type = request.form.get('supplier_type', '')
        products = request.form.get('products', '')
        supply_method = request.form.get('supply_method', '')
        notes = request.form.get('notes', '')
        
        supplier = Supplier(
            name=name,
            contact_person=contact_person,
            phone=phone,
            address=address,
            supplier_type=supplier_type,
            products=products,
            supply_method=supply_method,
            notes=notes
        )
        db.session.add(supplier)
        db.session.flush()  # 获取supplier.id
        
        # 处理图片上传
        if 'images' in request.files:
            files = request.files.getlist('images')
            for file in files:
                if file and file.filename and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    unique_filename = f"{uuid.uuid4().hex}_{filename}"
                    file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
                    file.save(file_path)
                    
                    supplier_image = SupplierImage(
                        supplier_id=supplier.id,
                        image_path=f"static/uploads/suppliers/{unique_filename}"
                    )
                    db.session.add(supplier_image)
        
        db.session.commit()
        flash('供应商添加成功！', 'success')
        return redirect(url_for('suppliers'))
    
    return render_template('add_supplier.html')

@app.route('/edit_supplier/<int:id>', methods=['GET', 'POST'])
@edit_required
def edit_supplier(id):
    supplier = Supplier.query.get_or_404(id)
    
    if request.method == 'POST':
        supplier.name = request.form['name']
        supplier.contact_person = request.form['contact_person']
        supplier.phone = request.form['phone']
        supplier.address = request.form['address']
        supplier.supplier_type = request.form.get('supplier_type', '')
        supplier.products = request.form.get('products', '')
        supplier.supply_method = request.form.get('supply_method', '')
        supplier.notes = request.form.get('notes', '')
        
        # 处理图片上传
        if 'images' in request.files:
            files = request.files.getlist('images')
            for file in files:
                if file and file.filename and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    unique_filename = f"{uuid.uuid4().hex}_{filename}"
                    file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
                    file.save(file_path)
                    
                    supplier_image = SupplierImage(
                        supplier_id=supplier.id,
                        image_path=f"static/uploads/suppliers/{unique_filename}"
                    )
                    db.session.add(supplier_image)
        
        db.session.commit()
        flash('供应商更新成功！', 'success')
        return redirect(url_for('suppliers'))
    
    return render_template('edit_supplier.html', supplier=supplier)

@app.route('/delete_supplier/<int:id>')
@edit_required
def delete_supplier(id):
    supplier = Supplier.query.get_or_404(id)
    
    # 检查是否有关联的交易或商品
    if supplier.transactions or supplier.product_list:
        flash('无法删除：该供应商下有关联的交易记录或商品', 'error')
    else:
        # 删除供应商图片
        for image in supplier.images:
            # image.image_path 已经包含了 static/uploads/suppliers/ 前缀
            if os.path.exists(image.image_path):
                os.remove(image.image_path)
            # 同时删除数据库中的图片记录
            db.session.delete(image)
        
        db.session.delete(supplier)
        db.session.commit()
        flash('供应商删除成功！', 'success')
    
    return redirect(url_for('suppliers'))

@app.route('/api/products')
@login_required
def api_products():
    """获取所有商品名称，用于供货品类选择"""
    products = Product.query.filter_by(is_active=True).all()
    product_names = [product.name for product in products]
    return jsonify(product_names)

@app.route('/api/product/<int:id>')
@login_required
def api_product_detail(id):
    """获取单个商品的详细信息"""
    product = Product.query.get_or_404(id)
    return jsonify({
        'id': product.id,
        'name': product.name,
        'code': f'PRD{product.id:03d}',  # 生成商品编号
        'category': product.category,
        'category_id': product.category,  # 暂时使用category字段
        'supplier_id': product.supplier_id,
        'cost': product.cost_price,
        'cost_price': product.cost_price,
        'price': product.selling_price,  # 前端使用price字段
        'selling_price': product.selling_price,
        'unit': product.unit,
        'stock': product.stock or 0,
        'min_stock': 10,  # 默认最低库存
        'max_stock': 1000,  # 默认最高库存
        'description': product.description or '',
        'image_path': product.image_path,
        'is_active': product.is_active,
        'created_at': product.created_at.strftime('%Y-%m-%d %H:%M:%S') if product.created_at else None
    })

@app.route('/api/product/<int:id>/stock', methods=['POST'])
@edit_required
def api_product_stock_operation(id):
    """商品库存操作"""
    product = Product.query.get_or_404(id)
    data = request.get_json()
    
    operation_type = data.get('operation_type')
    quantity = data.get('quantity', 0)
    remark = data.get('remark', '')
    
    try:
        if operation_type == 'in':
            # 入库
            product.stock += quantity
        elif operation_type == 'out':
            # 出库
            if product.stock < quantity:
                return jsonify({'success': False, 'message': '库存不足，无法出库'}), 400
            product.stock -= quantity
        elif operation_type == 'adjust':
            # 调整
            product.stock = quantity
        else:
            return jsonify({'success': False, 'message': '无效的操作类型'}), 400
        
        # 记录库存变动（这里可以添加库存变动记录表）
        # 暂时只更新商品库存
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '库存操作成功',
            'new_stock': product.stock
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'操作失败: {str(e)}'}), 500

@app.route('/api/supplier/<int:id>')
@login_required
def api_supplier_detail(id):
    """获取单个供应商的详细信息"""
    supplier = Supplier.query.get_or_404(id)
    return jsonify({
        'id': supplier.id,
        'name': supplier.name,
        'contact_person': supplier.contact_person,
        'phone': supplier.phone,
        'address': supplier.address,
        'supplier_type': supplier.supplier_type,
        'products': supplier.products,
        'supply_method': supplier.supply_method,
        'notes': supplier.notes,
        'is_active': supplier.is_active,
        'created_at': supplier.created_at.strftime('%Y-%m-%d %H:%M:%S') if supplier.created_at else None
    })

@app.route('/api/supplier/<int:id>', methods=['PUT'])
@edit_required
def api_update_supplier(id):
    """更新供应商信息"""
    supplier = Supplier.query.get_or_404(id)
    data = request.get_json()
    
    try:
        supplier.name = data.get('name', supplier.name)
        supplier.contact_person = data.get('contact_person', supplier.contact_person)
        supplier.phone = data.get('phone', supplier.phone)
        supplier.address = data.get('address', supplier.address)
        supplier.supplier_type = data.get('supplier_type', supplier.supplier_type)
        supplier.products = data.get('products', supplier.products)
        supplier.notes = data.get('notes', supplier.notes)
        
        db.session.commit()
        return jsonify({'message': '供应商更新成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'更新失败: {str(e)}'}), 500

@app.route('/api/supplier', methods=['POST'])
@edit_required
def api_create_supplier():
    """创建新供应商"""
    data = request.get_json()
    
    try:
        supplier = Supplier(
            name=data.get('name'),
            contact_person=data.get('contact_person'),
            phone=data.get('phone'),
            email=data.get('email'),
            address=data.get('address'),
            supplier_type=data.get('supplier_type'),
            products=data.get('products'),
            is_active=True
        )
        
        db.session.add(supplier)
        db.session.commit()
        return jsonify({'message': '供应商创建成功', 'id': supplier.id}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'创建失败: {str(e)}'}), 500

@app.route('/api/supplier/<int:id>/images')
@login_required
def api_supplier_images(id):
    """获取供应商的所有图片"""
    supplier = Supplier.query.get_or_404(id)
    images = [{'id': img.id, 'path': img.image_path, 'upload_time': img.upload_time.strftime('%Y-%m-%d %H:%M:%S')} for img in supplier.images]
    return jsonify(images)

@app.route('/api/supplier/<int:id>/images', methods=['POST'])
@edit_required
def api_upload_supplier_image(id):
    """上传供应商图片"""
    supplier = Supplier.query.get_or_404(id)
    
    # 支持单个文件上传（image字段）和多个文件上传（images字段）
    files = []
    if 'images' in request.files:
        files = request.files.getlist('images')
    elif 'image' in request.files:
        files = [request.files['image']]
    
    if not files:
        return jsonify({'error': '没有选择文件'}), 400
    
    uploaded_images = []
    
    for file in files:
        if file.filename == '':
            continue
            
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            unique_filename = f"{uuid.uuid4().hex}_{filename}"
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
            file.save(file_path)
            
            image_path = f"static/uploads/suppliers/{unique_filename}"
            supplier_image = SupplierImage(supplier_id=supplier.id, image_path=image_path)
            db.session.add(supplier_image)
            db.session.commit()
            
            uploaded_images.append({
                'id': supplier_image.id,
                'path': supplier_image.image_path,
                'upload_time': supplier_image.upload_time.strftime('%Y-%m-%d %H:%M:%S')
            })
    
    if uploaded_images:
        return jsonify({
            'success': True,
            'images': uploaded_images
        })
    else:
        return jsonify({'error': '没有有效的图片文件或不支持的文件格式'}), 400

@app.route('/api/supplier/<int:supplier_id>/images/<int:image_id>', methods=['DELETE'])
@edit_required
def api_delete_supplier_image(supplier_id, image_id):
    """删除供应商图片"""
    supplier_image = SupplierImage.query.get_or_404(image_id)
    
    if supplier_image.supplier_id != supplier_id:
        return jsonify({'error': '图片不属于该供应商'}), 400
    
    # 删除文件
    image_path = os.path.join('static', supplier_image.image_path)
    if os.path.exists(image_path):
        os.remove(image_path)
    
    db.session.delete(supplier_image)
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/products')
@login_required
def products():
    products = Product.query.all()
    suppliers = Supplier.query.all()
    categories = Category.query.all()
    return render_template('products.html', products=products, suppliers=suppliers, categories=categories)

@app.route('/add_product', methods=['GET', 'POST'])
@edit_required
def add_product():
    if request.method == 'POST':
        # 检查是否为AJAX请求
        is_ajax = 'X-Requested-With' in request.headers and request.headers['X-Requested-With'] == 'XMLHttpRequest'
        
        name = request.form['name']
        category = request.form.get('category')
        supplier_id = int(request.form['supplier']) if request.form.get('supplier') else None
        cost_price = float(request.form['cost_price']) if request.form.get('cost_price') else None
        selling_price = float(request.form['selling_price']) if request.form.get('selling_price') else None
        unit = request.form.get('unit', '件')
        description = request.form.get('description', '')
        is_active = request.form.get('is_active', '1') == '1'
        
        # 检查商品名称是否已存在
        if Product.query.filter_by(name=name).first():
            if is_ajax:
                return jsonify({'success': False, 'message': '商品名称已存在'}), 400
            flash('商品名称已存在', 'error')
            suppliers = Supplier.query.all()
            return render_template('add_product.html', suppliers=suppliers)
        
        # 处理图片上传
        image_path = None
        if 'image' in request.files:
            file = request.files['image']
            if file and file.filename != '' and allowed_file(file.filename):
                # 创建商品图片上传目录
                product_upload_folder = 'static/uploads/products'
                os.makedirs(product_upload_folder, exist_ok=True)
                
                # 生成唯一文件名
                filename = secure_filename(file.filename)
                unique_filename = f"{uuid.uuid4().hex}_{filename}"
                file_path = os.path.join(product_upload_folder, unique_filename)
                
                try:
                    file.save(file_path)
                    image_path = f"uploads/products/{unique_filename}"
                except Exception as e:
                    if is_ajax:
                        return jsonify({'success': False, 'message': f'图片上传失败: {str(e)}'}), 500
                    flash(f'图片上传失败: {str(e)}', 'error')
                    suppliers = Supplier.query.all()
                    return render_template('add_product.html', suppliers=suppliers)
        
        product = Product(
            name=name,
            category=category,
            supplier_id=supplier_id,
            cost_price=cost_price,
            selling_price=selling_price,
            unit=unit,
            description=description,
            image_path=image_path,
            is_active=is_active
        )
        db.session.add(product)
        db.session.commit()
        
        if is_ajax:
            return jsonify({'success': True})
        
        flash('商品添加成功！', 'success')
        return redirect(url_for('products'))
    
    suppliers = Supplier.query.all()
    categories = Category.query.all()  # 添加分类数据
    return render_template('add_product.html', suppliers=suppliers, categories=categories)

@app.route('/edit_product/<int:id>', methods=['GET', 'POST'])
@edit_required
def edit_product(id):
    product = Product.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            # 更新基本信息
            product.name = request.form['name']
            product.category = request.form['category']
            product.supplier_id = int(request.form['supplier']) if request.form.get('supplier') else None
            product.cost_price = float(request.form['cost_price']) if request.form['cost_price'] else None
            product.selling_price = float(request.form['selling_price']) if request.form['selling_price'] else None
            product.unit = request.form['unit']
            product.stock = int(request.form['stock']) if request.form['stock'] else 0
            product.description = request.form['description']
            product.is_active = request.form.get('is_active') == '1'
            
            # 检查是否需要删除图片
            remove_image = request.form.get('remove_image') == '1'
            print(f"Debug: remove_image = {remove_image}, product.image_path = {product.image_path}")
            if remove_image and product.image_path:
                # 删除旧图片文件
                old_image_path = os.path.join('static', product.image_path.lstrip('/'))
                print(f"Debug: 尝试删除图片文件: {old_image_path}")
                if os.path.exists(old_image_path):
                    os.remove(old_image_path)
                    print(f"Debug: 成功删除图片文件: {old_image_path}")
                else:
                    print(f"Debug: 图片文件不存在: {old_image_path}")
                # 清空数据库中的图片路径
                product.image_path = None
                print(f"Debug: 已清空数据库中的图片路径")
            
            # 处理图片上传
            if 'image' in request.files:
                file = request.files['image']
                if file and file.filename != '' and allowed_file(file.filename):
                    # 删除旧图片（如果存在且未被上面的删除逻辑处理）
                    if product.image_path and not remove_image:
                        old_image_path = os.path.join('static', product.image_path.lstrip('/'))
                        if os.path.exists(old_image_path):
                            os.remove(old_image_path)
                    
                    # 保存新图片
                    filename = secure_filename(file.filename)
                    unique_filename = f"{uuid.uuid4()}_{filename}"
                    
                    # 确保商品图片目录存在
                    product_upload_folder = 'static/uploads/products'
                    os.makedirs(product_upload_folder, exist_ok=True)
                    
                    file_path = os.path.join(product_upload_folder, unique_filename)
                    file.save(file_path)
                    product.image_path = f'uploads/products/{unique_filename}'
            
            db.session.commit()
            
            # 检查是否为AJAX请求
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': True})
            else:
                flash('商品信息更新成功！', 'success')
                return redirect(url_for('products'))
                
        except Exception as e:
            db.session.rollback()
            error_message = f'更新失败：{str(e)}'
            
            # 检查是否为AJAX请求
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': error_message})
            else:
                flash(error_message, 'error')
    
    suppliers = Supplier.query.all()
    return render_template('edit_product.html', product=product, suppliers=suppliers)

@app.route('/delete_product/<int:id>')
@edit_required
def delete_product(id):
    product = Product.query.get_or_404(id)
    
    # 检查是否为AJAX请求
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # 检查是否有关联的交易
        if product.transactions:
            return jsonify({'success': False, 'message': '无法删除：该商品下有关联的交易记录'})
        else:
            try:
                db.session.delete(product)
                db.session.commit()
                return jsonify({'success': True, 'message': '商品删除成功！'})
            except Exception as e:
                db.session.rollback()
                return jsonify({'success': False, 'message': f'删除失败：{str(e)}'})
    else:
        # 非AJAX请求，保持原有逻辑
        if product.transactions:
            flash('无法删除：该商品下有关联的交易记录', 'error')
        else:
            db.session.delete(product)
            db.session.commit()
            flash('商品删除成功！', 'success')
        
        return redirect(url_for('products'))

@app.route('/users')
@admin_required
def users():
    users = User.query.all()
    return render_template('users.html', users=users)

@app.route('/add_user', methods=['GET', 'POST'])
@admin_required
def add_user():
    if request.method == 'POST':
        # 检查是否为AJAX请求
        is_ajax = request.headers.get('Content-Type') == 'multipart/form-data' or request.is_json or 'XMLHttpRequest' in request.headers.get('X-Requested-With', '')
        
        username = request.form['username']
        email = request.form['email']
        password = request.form['password']
        role = request.form['role']
        is_active = request.form.get('is_active', '1') == '1'
        
        # 检查用户名是否已存在
        if User.query.filter_by(username=username).first():
            if is_ajax:
                return jsonify({'success': False, 'message': '用户名已存在'}), 400
            flash('用户名已存在', 'error')
            return render_template('add_user.html')
        
        # 检查邮箱是否已存在
        if User.query.filter_by(email=email).first():
            if is_ajax:
                return jsonify({'success': False, 'message': '邮箱已存在'}), 400
            flash('邮箱已存在', 'error')
            return render_template('add_user.html')
        
        user = User(
            username=username,
            email=email,
            password_hash=generate_password_hash(password),
            role=role,
            is_active=is_active
        )
        db.session.add(user)
        db.session.commit()
        
        if is_ajax:
            return jsonify({'success': True, 'message': '用户添加成功！'})
        
        flash('用户添加成功！', 'success')
        return redirect(url_for('users'))
    
    return render_template('add_user.html')

@app.route('/edit_user/<int:id>', methods=['GET', 'POST'])
@admin_required
def edit_user(id):
    user = User.query.get_or_404(id)
    
    if request.method == 'POST':
        # 检查是否为AJAX请求
        is_ajax = request.headers.get('Content-Type') == 'multipart/form-data' or request.is_json or 'XMLHttpRequest' in request.headers.get('X-Requested-With', '')
        
        username = request.form['username']
        email = request.form['email']
        role = request.form['role']
        is_active = 'is_active' in request.form
        
        # 检查用户名是否已存在（排除当前用户）
        existing_user = User.query.filter_by(username=username).first()
        if existing_user and existing_user.id != user.id:
            if is_ajax:
                return jsonify({'success': False, 'message': '用户名已存在'}), 400
            flash('用户名已存在', 'error')
            return render_template('edit_user.html', user=user)
        
        # 检查邮箱是否已存在（排除当前用户）
        existing_email = User.query.filter_by(email=email).first()
        if existing_email and existing_email.id != user.id:
            if is_ajax:
                return jsonify({'success': False, 'message': '邮箱已存在'}), 400
            flash('邮箱已存在', 'error')
            return render_template('edit_user.html', user=user)
        
        user.username = username
        user.email = email
        user.role = role
        user.is_active = is_active
        
        # 如果提供了新密码，则更新密码
        new_password = request.form.get('password')
        if new_password:
            user.password_hash = generate_password_hash(new_password)
        
        db.session.commit()
        
        if is_ajax:
            return jsonify({'success': True, 'message': '用户更新成功！'})
        
        flash('用户更新成功！', 'success')
        return redirect(url_for('users'))
    
    return render_template('edit_user.html', user=user)

@app.route('/delete_user/<int:id>')
@admin_required
def delete_user(id):
    user = User.query.get_or_404(id)
    
    # 不能删除自己
    if user.id == current_user.id:
        flash('不能删除自己的账户', 'error')
        return redirect(url_for('users'))
    
    # 不能删除最后一个管理员
    if user.role == 'admin':
        admin_count = User.query.filter_by(role='admin').count()
        if admin_count <= 1:
            flash('不能删除最后一个管理员账户', 'error')
            return redirect(url_for('users'))
    
    db.session.delete(user)
    db.session.commit()
    flash('用户删除成功！', 'success')
    
    return redirect(url_for('users'))

@app.route('/statistics')
@login_required
def statistics():
    """财务分析页面 - 增强版本"""
    try:
        # 获取筛选参数
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # 获取当前日期
        today = datetime.now().date()
        month_start = today.replace(day=1)
        
        # 设置默认时间范围（最近12个月）
        if not start_date:
            default_start = today - timedelta(days=365)
            start_date = default_start.strftime('%Y-%m-%d')
        if not end_date:
            end_date = today.strftime('%Y-%m-%d')
            
        # 转换日期
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            start_date_obj = today - timedelta(days=365)
            end_date_obj = today
        
        # 基础财务数据（基于选定时间范围）
        total_income = db.session.query(db.func.sum(Transaction.amount)).filter(
            Transaction.type == 'income',
            Transaction.date >= start_date_obj,
            Transaction.date <= end_date_obj
        ).scalar() or 0
        
        total_expense = db.session.query(db.func.sum(Transaction.amount)).filter(
            Transaction.type == 'expense',
            Transaction.date >= start_date_obj,
            Transaction.date <= end_date_obj
        ).scalar() or 0
        
        net_assets = total_income - total_expense
        liability_ratio = (total_expense / total_income * 100) if total_income > 0 else 0
        
        # 财务健康度评分
        if net_assets > 0:
            health_score = min(100, max(0, int(100 - liability_ratio)))
        else:
            health_score = 0
        
        # 获取月度统计数据
        monthly_stats = db.session.query(
            db.func.date_format(Transaction.date, '%Y-%m').label('month'),
            Transaction.type,
            db.func.sum(Transaction.amount).label('total')
        ).filter(
            Transaction.date >= start_date_obj,
            Transaction.date <= end_date_obj
        ).group_by(
            db.func.date_format(Transaction.date, '%Y-%m'),
            Transaction.type
        ).order_by('month').all()
        
        # 处理趋势数据
        month_data = defaultdict(lambda: {'income': 0, 'expense': 0})
        for stat in monthly_stats:
            month_data[stat.month][stat.type] = stat.total
        
        trend_months = []
        trend_income = []
        trend_expense = []
        trend_profit = []
        monthly_data = []
        
        # 计算环比和同比增长
        sorted_months = sorted(month_data.keys())
        for i, month in enumerate(sorted_months):
            income = month_data[month]['income']
            expense = month_data[month]['expense']
            profit = income - expense
            
            # 环比增长（与上月比较）
            mom_growth = 0
            if i > 0:
                prev_month = sorted_months[i-1]
                prev_profit = month_data[prev_month]['income'] - month_data[prev_month]['expense']
                if prev_profit != 0:
                    mom_growth = ((profit - prev_profit) / abs(prev_profit)) * 100
            
            # 同比增长（与去年同月比较）
            yoy_growth = 0
            year_ago_month = f"{int(month[:4])-1}-{month[5:]}"
            if year_ago_month in month_data:
                year_ago_profit = month_data[year_ago_month]['income'] - month_data[year_ago_month]['expense']
                if year_ago_profit != 0:
                    yoy_growth = ((profit - year_ago_profit) / abs(year_ago_profit)) * 100
            
            trend_months.append(month)
            trend_income.append(income)
            trend_expense.append(expense)
            trend_profit.append(profit)
            
            monthly_data.append({
                'date': month,
                'income': income,
                'expense': expense,
                'profit': profit,
                'mom_growth': mom_growth,
                'yoy_growth': yoy_growth
            })
        
        # 收支分类数据（基于选定时间范围）
        income_categories = db.session.query(
            Category.name,
            db.func.sum(Transaction.amount).label('amount')
        ).join(Transaction).filter(
            Transaction.type == 'income',
            Transaction.date >= start_date_obj,
            Transaction.date <= end_date_obj
        ).group_by(Category.name).all()
        
        expense_categories = db.session.query(
            Category.name,
            db.func.sum(Transaction.amount).label('amount')
        ).join(Transaction).filter(
            Transaction.type == 'expense',
            Transaction.date >= start_date_obj,
            Transaction.date <= end_date_obj
        ).group_by(Category.name).all()
        
        # 饼图数据
        income_colors = ['#57B5E7', '#8DD3C7', '#9333EA', '#FB923C', '#F59E0B']
        expense_colors = ['#FBBF72', '#FC8D62', '#EAB308', '#EF4444', '#8B5CF6']
        
        income_pie_data = [
            {
                'value': cat.amount, 
                'name': cat.name, 
                'itemStyle': {'color': income_colors[i % len(income_colors)]}
            }
            for i, cat in enumerate(income_categories)
        ]
        
        expense_pie_data = [
            {
                'value': cat.amount, 
                'name': cat.name, 
                'itemStyle': {'color': expense_colors[i % len(expense_colors)]}
            }
            for i, cat in enumerate(expense_categories)
        ]
        
        # 现金流瀑布图数据（简化版）
        # 本月数据
        current_month_income = db.session.query(db.func.sum(Transaction.amount)).filter(
            Transaction.type == 'income',
            Transaction.date >= month_start,
            Transaction.date <= today
        ).scalar() or 0
        
        current_month_expense = db.session.query(db.func.sum(Transaction.amount)).filter(
            Transaction.type == 'expense',
            Transaction.date >= month_start,
            Transaction.date <= today
        ).scalar() or 0
        
        # 上月结余（简化计算）
        last_month_end = month_start - timedelta(days=1)
        last_month_start = last_month_end.replace(day=1)
        
        last_month_income = db.session.query(db.func.sum(Transaction.amount)).filter(
            Transaction.type == 'income',
            Transaction.date >= last_month_start,
            Transaction.date <= last_month_end
        ).scalar() or 0
        
        last_month_expense = db.session.query(db.func.sum(Transaction.amount)).filter(
            Transaction.type == 'expense',
            Transaction.date >= last_month_start,
            Transaction.date <= last_month_end
        ).scalar() or 0
        
        initial_cash = last_month_income - last_month_expense
        operating_cash = current_month_income
        investing_cash = -current_month_expense * 0.3  # 假设30%为投资支出
        financing_cash = -current_month_expense * 0.7  # 假设70%为运营支出
        final_cash = initial_cash + operating_cash + investing_cash + financing_cash
        
        cashflow_data = [
            {'value': float(initial_cash), 'itemStyle': {'color': '#57B5E7'}},
            {'value': float(operating_cash), 'itemStyle': {'color': '#8DD3C7'}},
            {'value': float(investing_cash), 'itemStyle': {'color': '#FC8D62'}},
            {'value': float(financing_cash), 'itemStyle': {'color': '#FBBF72'}},
            {'value': float(final_cash), 'itemStyle': {'color': '#57B5E7'}}
        ]
        
        # 增长率数据（简化版）
        assets_growth = 0
        liabilities_growth = 0
        net_assets_growth = 0
        liability_ratio_change = 0
        
        return render_template('statistics.html',
                             total_assets=total_income,
                             total_liabilities=total_expense,
                             net_assets=net_assets,
                             liability_ratio=liability_ratio,
                             health_score=health_score,
                             trend_months=trend_months,
                             trend_income=trend_income,
                             trend_expense=trend_expense,
                             trend_profit=trend_profit,
                             income_categories=income_categories,
                             expense_categories=expense_categories,
                             income_pie_data=income_pie_data,
                             expense_pie_data=expense_pie_data,
                             cashflow_data=cashflow_data,
                             monthly_data=monthly_data,
                             assets_growth=assets_growth,
                             liabilities_growth=liabilities_growth,
                             net_assets_growth=net_assets_growth,
                             liability_ratio_change=liability_ratio_change,
                             start_date=start_date,
                             end_date=end_date)
    
    except Exception as e:
        flash(f'财务分析数据加载失败: {str(e)}', 'error')
        return render_template('statistics.html',
                             total_assets=0,
                             total_liabilities=0,
                             net_assets=0,
                             liability_ratio=0,
                             health_score=0,
                             trend_months=[],
                             trend_income=[],
                             trend_expense=[],
                             trend_profit=[],
                             income_categories=[],
                             expense_categories=[],
                             income_pie_data=[],
                             expense_pie_data=[],
                             cashflow_data=[],
                             monthly_data=[],
                             assets_growth=0,
                             liabilities_growth=0,
                             net_assets_growth=0,
                             liability_ratio_change=0)

@app.route('/export_statistics')
@login_required
def export_statistics():
    """导出财务分析数据"""
    try:
        # 获取筛选参数
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        export_format = request.args.get('format', 'csv')
        
        # 获取当前日期
        today = datetime.now().date()
        
        # 设置默认时间范围
        if not start_date:
            default_start = today - timedelta(days=365)
            start_date = default_start.strftime('%Y-%m-%d')
        if not end_date:
            end_date = today.strftime('%Y-%m-%d')
            
        # 转换日期
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            start_date_obj = today - timedelta(days=365)
            end_date_obj = today
        
        # 获取月度统计数据
        monthly_stats = db.session.query(
            db.func.date_format(Transaction.date, '%Y-%m').label('month'),
            Transaction.type,
            db.func.sum(Transaction.amount).label('total')
        ).filter(
            Transaction.date >= start_date_obj,
            Transaction.date <= end_date_obj
        ).group_by(
            db.func.date_format(Transaction.date, '%Y-%m'),
            Transaction.type
        ).order_by('month').all()
        
        # 处理数据
        month_data = defaultdict(lambda: {'income': 0, 'expense': 0})
        for stat in monthly_stats:
            month_data[stat.month][stat.type] = stat.total
        
        # 计算增长率
        export_data = []
        sorted_months = sorted(month_data.keys())
        for i, month in enumerate(sorted_months):
            income = month_data[month]['income']
            expense = month_data[month]['expense']
            profit = income - expense
            
            # 环比增长
            mom_growth = 0
            if i > 0:
                prev_month = sorted_months[i-1]
                prev_profit = month_data[prev_month]['income'] - month_data[prev_month]['expense']
                if prev_profit != 0:
                    mom_growth = ((profit - prev_profit) / abs(prev_profit)) * 100
            
            # 同比增长
            yoy_growth = 0
            year_ago_month = f"{int(month[:4])-1}-{month[5:]}"
            if year_ago_month in month_data:
                year_ago_profit = month_data[year_ago_month]['income'] - month_data[year_ago_month]['expense']
                if year_ago_profit != 0:
                    yoy_growth = ((profit - year_ago_profit) / abs(year_ago_profit)) * 100
            
            export_data.append({
                'month': month,
                'income': income,
                'expense': expense,
                'profit': profit,
                'mom_growth': mom_growth,
                'yoy_growth': yoy_growth
            })
        
        if not export_data:
            flash('没有符合条件的财务数据可以导出', 'warning')
            return redirect(url_for('statistics'))
        
        if export_format == 'csv':
            # 创建CSV文件
            output = io.StringIO()
            writer = csv.writer(output)
            
            # 写入表头
            writer.writerow(['月份', '收入', '支出', '利润', '环比增长(%)', '同比增长(%)'])
            
            # 写入数据
            for data in export_data:
                writer.writerow([
                    data['month'],
                    f"{data['income']:.2f}",
                    f"{data['expense']:.2f}",
                    f"{data['profit']:.2f}",
                    f"{data['mom_growth']:.2f}",
                    f"{data['yoy_growth']:.2f}"
                ])
            
            # 创建响应
            response = make_response(output.getvalue())
            response.headers['Content-Type'] = 'text/csv; charset=utf-8'
            response.headers['X-Content-Type-Options'] = 'nosniff'
            response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            response.headers['Pragma'] = 'no-cache'
            response.headers['Expires'] = '0'
            
            # 生成文件名
            filename = f'financial_analysis_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
            filename_utf8 = f'财务分析_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
            response.headers['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'\'{filename_utf8.encode("utf-8").decode("latin1")}"'
            
            return response
            
        elif export_format == 'excel':
            if not EXCEL_AVAILABLE:
                flash('Excel导出功能不可用，请安装openpyxl库或选择CSV格式导出', 'error')
                return redirect(url_for('statistics'))
                
            # 创建Excel文件
            output = io.BytesIO()
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            
            workbook = Workbook()
            ws = workbook.active
            ws.title = "财务分析数据"
            
            # 设置表头
            headers = ['月份', '收入', '支出', '利润', '环比增长(%)', '同比增长(%)']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # 写入数据
            for row, data in enumerate(export_data, 2):
                ws.cell(row=row, column=1, value=data['month'])
                ws.cell(row=row, column=2, value=data['income'])
                ws.cell(row=row, column=3, value=data['expense'])
                ws.cell(row=row, column=4, value=data['profit'])
                ws.cell(row=row, column=5, value=data['mom_growth'])
                ws.cell(row=row, column=6, value=data['yoy_growth'])
            
            # 设置列宽
            column_widths = [12, 15, 15, 15, 15, 15]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(col)].width = width
            
            workbook.save(output)
            output.seek(0)
            
            # 创建响应
            response = make_response(output.getvalue())
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            
            # 生成文件名
            filename = f'financial_analysis_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            filename_utf8 = f'财务分析_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            response.headers['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'\'{filename_utf8.encode("utf-8").decode("latin1")}"'
            
            # 添加安全头部减少浏览器警告
            response.headers['X-Content-Type-Options'] = 'nosniff'
            response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            response.headers['Pragma'] = 'no-cache'
            response.headers['Expires'] = '0'
            
            return response
            
    except Exception as e:
        flash(f'导出失败: {str(e)}', 'error')
        return redirect(url_for('statistics'))

@app.route('/export_pdf_report')
@login_required
def export_pdf_report():
    """导出PDF财务分析报告"""
    try:
        if not PDF_AVAILABLE:
            flash('PDF生成功能不可用，请安装reportlab库', 'error')
            return redirect(url_for('statistics'))
        
        # 获取筛选参数
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # 获取当前日期
        today = datetime.now().date()
        
        # 设置默认时间范围
        if not start_date:
            default_start = today - timedelta(days=365)
            start_date = default_start.strftime('%Y-%m-%d')
        if not end_date:
            end_date = today.strftime('%Y-%m-%d')
            
        # 转换日期
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            start_date_obj = today - timedelta(days=365)
            end_date_obj = today
        
        # 获取财务数据
        monthly_stats = db.session.query(
            db.func.date_format(Transaction.date, '%Y-%m').label('month'),
            Transaction.type,
            db.func.sum(Transaction.amount).label('total')
        ).filter(
            Transaction.date >= start_date_obj,
            Transaction.date <= end_date_obj
        ).group_by(
            db.func.date_format(Transaction.date, '%Y-%m'),
            Transaction.type
        ).order_by('month').all()
        
        # 处理数据
        month_data = defaultdict(lambda: {'income': 0, 'expense': 0})
        for stat in monthly_stats:
            month_data[stat.month][stat.type] = stat.total
        
        # 计算关键指标
        total_income = sum(data['income'] for data in month_data.values())
        total_expense = sum(data['expense'] for data in month_data.values())
        net_profit = total_income - total_expense
        
        # 获取交易数据用于分析
        transactions = Transaction.query.filter(
            Transaction.date >= start_date_obj,
            Transaction.date <= end_date_obj
        ).all()
        
        # 注册中文字体
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        
        # 尝试注册系统中文字体
        chinese_font = 'Helvetica'  # 默认字体
        try:
            # Windows系统字体路径
            font_paths = [
                'C:/Windows/Fonts/simsun.ttc',  # 宋体
                'C:/Windows/Fonts/simhei.ttf',  # 黑体
                'C:/Windows/Fonts/msyh.ttc',    # 微软雅黑
                'C:/Windows/Fonts/simkai.ttf'   # 楷体
            ]
            
            for font_path in font_paths:
                if os.path.exists(font_path):
                    try:
                        pdfmetrics.registerFont(TTFont('ChineseFont', font_path))
                        chinese_font = 'ChineseFont'
                        break
                    except:
                        continue
        except Exception as e:
            print(f"字体注册失败: {e}")
        
        # 创建PDF
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4, topMargin=1*inch, bottomMargin=1*inch)
        
        # 获取样式
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontName=chinese_font,
            fontSize=24,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.darkblue
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontName=chinese_font,
            fontSize=16,
            spaceAfter=12,
            textColor=colors.darkblue
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontName=chinese_font,
            fontSize=10
        )
        
        # 构建PDF内容
        story = []
        
        # 标题
        story.append(Paragraph('七彩果坊财务分析报告', title_style))
        story.append(Paragraph(f'报告期间: {start_date} 至 {end_date}', normal_style))
        story.append(Paragraph(f'生成时间: {datetime.now().strftime("%Y年%m月%d日 %H:%M")}', normal_style))
        story.append(Spacer(1, 20))
        
        # 执行摘要
        story.append(Paragraph('执行摘要', heading_style))
        summary_data = [
            ['指标', '金额', '说明'],
            ['总收入', f'¥{total_income:,.2f}', '报告期内总收入'],
            ['总支出', f'¥{total_expense:,.2f}', '报告期内总支出'],
            ['净利润', f'¥{net_profit:,.2f}', '收入减去支出的净额'],
            ['利润率', f'{(net_profit/total_income*100) if total_income > 0 else 0:.2f}%', '净利润占总收入的比例']
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 1.5*inch, 2.5*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), chinese_font),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # 添加图表分析部分
        story.append(Paragraph('图表分析', heading_style))
        
        # 创建收入支出趋势图
        if month_data:
            # 准备趋势图数据
            months = sorted(month_data.keys())
            income_data = [month_data[month]['income'] for month in months]
            expense_data = [month_data[month]['expense'] for month in months]
            
            # 创建趋势图
            drawing = Drawing(400, 200)
            chart = HorizontalLineChart()
            chart.x = 50
            chart.y = 50
            chart.height = 125
            chart.width = 300
            chart.data = [income_data, expense_data]
            chart.categoryAxis.categoryNames = [m.replace('-', '/') for m in months]
            chart.valueAxis.valueMin = 0
            chart.valueAxis.valueMax = max(max(income_data) if income_data else [0], max(expense_data) if expense_data else [0]) * 1.1
            chart.lines[0].strokeColor = colors.green
            chart.lines[1].strokeColor = colors.red
            chart.lines[0].strokeWidth = 2
            chart.lines[1].strokeWidth = 2
            drawing.add(chart)
            
            # 添加图例
            from reportlab.graphics.shapes import String
            drawing.add(String(50, 30, '绿线: 收入', fontSize=10, fillColor=colors.green))
            drawing.add(String(150, 30, '红线: 支出', fontSize=10, fillColor=colors.red))
            
            story.append(drawing)
            story.append(Spacer(1, 10))
            
            # 趋势分析文字
            trend_analysis = []
            if len(months) >= 2:
                recent_income = income_data[-1] if income_data else 0
                previous_income = income_data[-2] if len(income_data) >= 2 else 0
                recent_expense = expense_data[-1] if expense_data else 0
                previous_expense = expense_data[-2] if len(expense_data) >= 2 else 0
                
                income_change = ((recent_income - previous_income) / previous_income * 100) if previous_income > 0 else 0
                expense_change = ((recent_expense - previous_expense) / previous_expense * 100) if previous_expense > 0 else 0
                
                trend_analysis.append(f'• 最近一个月收入{"增长" if income_change > 0 else "下降"}{abs(income_change):.1f}%')
                trend_analysis.append(f'• 最近一个月支出{"增长" if expense_change > 0 else "下降"}{abs(expense_change):.1f}%')
                
                if income_change > expense_change:
                    trend_analysis.append('• 收入增长速度超过支出增长，财务状况向好')
                elif expense_change > income_change:
                    trend_analysis.append('• 支出增长速度超过收入增长，需要关注成本控制')
            
            for analysis in trend_analysis:
                story.append(Paragraph(analysis, normal_style))
            
            story.append(Spacer(1, 20))
        
        # 获取分类数据用于饼图
        category_stats = db.session.query(
            Category.name,
            Transaction.type,
            db.func.sum(Transaction.amount).label('total')
        ).join(
            Transaction, Category.id == Transaction.category_id
        ).filter(
            Transaction.date >= start_date_obj,
            Transaction.date <= end_date_obj
        ).group_by(
            Category.name, Transaction.type
        ).all()
        
        # 处理分类数据
        income_categories = {}
        expense_categories = {}
        for stat in category_stats:
            if stat.type == 'income':
                income_categories[stat.name] = stat.total
            else:
                expense_categories[stat.name] = stat.total
        
        # 创建支出分类饼图（如果有数据）
        if expense_categories:
            story.append(Paragraph('支出分类分析', heading_style))
            
            # 创建饼图
            drawing = Drawing(500, 300)
            pie = Pie()
            pie.x = 50
            pie.y = 50
            pie.width = 180
            pie.height = 180
            
            # 准备饼图数据
            categories = list(expense_categories.keys())
            values = list(expense_categories.values())
            pie.data = values
            
            # 设置颜色
            colors_list = [colors.red, colors.orange, colors.yellow, colors.green, colors.blue, colors.purple]
            pie.slices.strokeColor = colors.white
            for i, color in enumerate(colors_list[:len(values)]):
                pie.slices[i].fillColor = color
            
            # 设置字体
            pie.slices.fontName = chinese_font
            pie.slices.fontSize = 10
            pie.slices.labelRadius = 1.2
            
            # 创建图例
            legend_x = 280
            legend_y = 200
            for i, (category, value) in enumerate(zip(categories, values)):
                # 颜色块
                rect = Rect(legend_x, legend_y - i * 20, 10, 10)
                rect.fillColor = colors_list[i % len(colors_list)]
                rect.strokeColor = colors.black
                drawing.add(rect)
                
                # 标签文字
                percentage = (value / sum(values)) * 100
                label_text = f'{category}: {percentage:.1f}%'
                label = String(legend_x + 15, legend_y - i * 20 + 2, label_text)
                label.fontName = chinese_font
                label.fontSize = 9
                drawing.add(label)
            
            # 移除饼图自带的标签以避免重叠
            pie.labels = None
            
            drawing.add(pie)
            story.append(drawing)
            story.append(Spacer(1, 10))
            
            # 分类分析文字
            total_expense_cat = sum(expense_categories.values())
            category_analysis = []
            sorted_categories = sorted(expense_categories.items(), key=lambda x: x[1], reverse=True)
            
            if sorted_categories:
                top_category = sorted_categories[0]
                category_analysis.append(f'• 最大支出类别：{top_category[0]}，占总支出的{(top_category[1]/total_expense_cat*100):.1f}%')
                
                if len(sorted_categories) >= 3:
                    top3_total = sum([cat[1] for cat in sorted_categories[:3]])
                    category_analysis.append(f'• 前三大支出类别占总支出的{(top3_total/total_expense_cat*100):.1f}%')
                
                category_analysis.append('• 建议重点关注主要支出类别的成本控制')
            
            for analysis in category_analysis:
                story.append(Paragraph(analysis, normal_style))
            
            story.append(Spacer(1, 20))
        
        # 月度收支明细
        story.append(Paragraph('月度收支明细', heading_style))
        detail_data = [['月份', '收入', '支出', '净利润', '利润率']]
        
        for month in sorted(month_data.keys()):
            data = month_data[month]
            monthly_profit = data['income'] - data['expense']
            profit_rate = (monthly_profit / data['income'] * 100) if data['income'] > 0 else 0
            detail_data.append([
                month,
                f'¥{data["income"]:,.2f}',
                f'¥{data["expense"]:,.2f}',
                f'¥{monthly_profit:,.2f}',
                f'{profit_rate:.2f}%'
            ])
        
        detail_table = Table(detail_data, colWidths=[1.2*inch, 1.2*inch, 1.2*inch, 1.2*inch, 1.2*inch])
        detail_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), chinese_font),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(detail_table)
        story.append(Spacer(1, 20))
        
        # 收入支出对比分析图表
        story.append(Paragraph('收入支出对比分析', heading_style))
        story.append(Spacer(1, 10))
        
        # 创建收入支出对比柱状图
        if month_data:
            drawing = Drawing(500, 300)
            chart = VerticalBarChart()
            chart.x = 50
            chart.y = 50
            chart.height = 200
            chart.width = 400
            
            # 准备数据
            months = sorted(month_data.keys())[-6:]  # 最近6个月
            income_values = [month_data[m]['income'] for m in months]
            expense_values = [month_data[m]['expense'] for m in months]
            
            chart.data = [income_values, expense_values]
            chart.categoryAxis.categoryNames = [m.split('-')[1] + '月' for m in months]  # 只显示月份
            chart.bars[0].fillColor = colors.green
            chart.bars[1].fillColor = colors.red
            chart.valueAxis.valueMin = 0
            chart.valueAxis.valueMax = max(max(income_values, default=0), max(expense_values, default=0)) * 1.1
            
            # 添加图例
            chart.categoryAxis.labels.fontName = chinese_font
            chart.valueAxis.labels.fontName = chinese_font
            
            drawing.add(chart)
            story.append(drawing)
            story.append(Spacer(1, 10))
            
            # 收入支出对比分析文字
            comparison_analysis = []
            avg_income = sum(income_values) / len(income_values) if income_values else 0
            avg_expense = sum(expense_values) / len(expense_values) if expense_values else 0
            
            comparison_analysis.append(f'• 近{len(months)}个月平均收入为 {avg_income:,.2f} 元，平均支出为 {avg_expense:,.2f} 元。')
            
            if avg_income > avg_expense:
                comparison_analysis.append(f'• 收入高于支出 {avg_income - avg_expense:,.2f} 元，经营状况良好。')
            elif avg_income < avg_expense:
                comparison_analysis.append(f'• 支出高于收入 {avg_expense - avg_income:,.2f} 元，需要控制成本或增加收入。')
            else:
                comparison_analysis.append('• 收支基本平衡。')
            
            # 分析收入趋势
            if len(income_values) >= 3:
                recent_income_trend = income_values[-1] - income_values[-3]
                if recent_income_trend > 0:
                    comparison_analysis.append(f'• 收入呈上升趋势，较前期增长 {recent_income_trend:,.2f} 元。')
                elif recent_income_trend < 0:
                    comparison_analysis.append(f'• 收入呈下降趋势，较前期减少 {abs(recent_income_trend):,.2f} 元。')
                else:
                    comparison_analysis.append('• 收入保持稳定。')
            
            for analysis in comparison_analysis:
                story.append(Paragraph(analysis, normal_style))
        
        story.append(Spacer(1, 20))
        
        # 月度净利润趋势分析
        story.append(Paragraph('月度净利润趋势分析', heading_style))
        story.append(Spacer(1, 10))
        
        if month_data:
            # 创建净利润趋势线图
            drawing = Drawing(500, 300)
            chart = HorizontalLineChart()
            chart.x = 50
            chart.y = 50
            chart.height = 200
            chart.width = 400
            
            # 准备净利润数据
            profit_values = [month_data[m]['income'] - month_data[m]['expense'] for m in months]
            chart.data = [profit_values]
            chart.categoryAxis.categoryNames = [m.split('-')[1] + '月' for m in months]
            chart.lines[0].strokeColor = colors.blue
            chart.lines[0].strokeWidth = 2
            
            # 设置Y轴范围
            min_profit = min(profit_values) if profit_values else 0
            max_profit = max(profit_values) if profit_values else 0
            chart.valueAxis.valueMin = min_profit * 1.1 if min_profit < 0 else min_profit * 0.9
            chart.valueAxis.valueMax = max_profit * 1.1 if max_profit > 0 else max_profit * 0.9
            
            # 设置字体
            chart.categoryAxis.labels.fontName = chinese_font
            chart.valueAxis.labels.fontName = chinese_font
            
            drawing.add(chart)
            story.append(drawing)
            story.append(Spacer(1, 10))
            
            # 净利润趋势分析文字
            profit_analysis = []
            avg_profit = sum(profit_values) / len(profit_values) if profit_values else 0
            profit_analysis.append(f'• 近{len(months)}个月平均净利润为 {avg_profit:,.2f} 元。')
            
            # 分析利润波动
            if len(profit_values) >= 2:
                profit_volatility = max(profit_values) - min(profit_values)
                profit_analysis.append(f'• 净利润波动范围为 {profit_volatility:,.2f} 元。')
                
                if profit_volatility > abs(avg_profit) * 0.5:
                    profit_analysis.append('• 利润波动较大，建议加强经营稳定性。')
                else:
                    profit_analysis.append('• 利润波动相对稳定。')
            
            # 分析最近趋势
            if len(profit_values) >= 3:
                recent_trend = profit_values[-1] - profit_values[-3]
                if recent_trend > 0:
                    profit_analysis.append(f'• 最近净利润呈上升趋势，增长 {recent_trend:,.2f} 元。')
                elif recent_trend < 0:
                    profit_analysis.append(f'• 最近净利润呈下降趋势，减少 {abs(recent_trend):,.2f} 元。')
                else:
                    profit_analysis.append('• 最近净利润保持稳定。')
            
            for analysis in profit_analysis:
                story.append(Paragraph(analysis, normal_style))
        
        story.append(Spacer(1, 20))
        
        # 支出类别分布分析
        story.append(Paragraph('支出类别分布分析', heading_style))
        story.append(Spacer(1, 10))
        
        # 获取支出类别数据
        expense_categories = {}
        for transaction in transactions:
            if transaction.type == 'expense':
                category_name = transaction.category.name
                expense_categories[category_name] = expense_categories.get(category_name, 0) + transaction.amount
        
        if expense_categories:
            # 创建饼图
            drawing = Drawing(500, 300)
            pie = Pie()
            pie.x = 50
            pie.y = 50
            pie.width = 180
            pie.height = 180
            
            # 准备饼图数据
            categories = list(expense_categories.keys())
            amounts = list(expense_categories.values())
            
            pie.data = amounts
            pie.labels = categories
            
            # 设置颜色
            colors_list = [colors.red, colors.blue, colors.green, colors.orange, colors.purple, colors.brown, colors.pink, colors.gray]
            for i in range(len(amounts)):
                pie.slices[i].fillColor = colors_list[i % len(colors_list)]
            
            # 设置标签字体和位置
            pie.slices.fontName = chinese_font
            pie.slices.fontSize = 10
            pie.slices.labelRadius = 1.2
            pie.slices.popout = 0
            
            # 创建图例
            legend_x = 280
            legend_y = 200
            for i, (category, amount) in enumerate(zip(categories, amounts)):
                # 颜色块
                rect = Rect(legend_x, legend_y - i * 20, 10, 10)
                rect.fillColor = colors_list[i % len(colors_list)]
                rect.strokeColor = colors.black
                drawing.add(rect)
                
                # 标签文字
                percentage = (amount / sum(amounts)) * 100
                label_text = f'{category}: {percentage:.1f}%'
                label = String(legend_x + 15, legend_y - i * 20 + 2, label_text)
                label.fontName = chinese_font
                label.fontSize = 9
                drawing.add(label)
            
            # 移除饼图自带的标签以避免重叠
            pie.labels = None
            
            drawing.add(pie)
            story.append(drawing)
            story.append(Spacer(1, 10))
            
            # 支出类别分析文字
            category_analysis = []
            total_expense_amount = sum(amounts)
            
            # 找出最大支出类别
            max_category = max(expense_categories, key=expense_categories.get)
            max_amount = expense_categories[max_category]
            max_percentage = (max_amount / total_expense_amount) * 100
            
            category_analysis.append(f'• 最大支出类别为"{max_category}"，金额为 {max_amount:,.2f} 元，占总支出的 {max_percentage:.1f}%。')
            
            # 分析支出集中度
            if max_percentage > 50:
                category_analysis.append('• 支出过于集中在单一类别，建议分散支出风险。')
            elif max_percentage > 30:
                category_analysis.append('• 支出相对集中，建议关注主要支出类别的成本控制。')
            else:
                category_analysis.append('• 支出分布相对均衡。')
            
            # 列出前三大支出类别
            sorted_categories = sorted(expense_categories.items(), key=lambda x: x[1], reverse=True)
            top_3_categories = sorted_categories[:3]
            
            category_analysis.append('• 前三大支出类别分别为：')
            for i, (cat, amount) in enumerate(top_3_categories, 1):
                percentage = (amount / total_expense_amount) * 100
                category_analysis.append(f'  {i}. {cat}：{amount:,.2f} 元 ({percentage:.1f}%)')
            
            # 成本控制建议
            if max_percentage > 40:
                category_analysis.append(f'• 建议重点关注"{max_category}"类别的成本控制，寻找节约空间。')
            
            for analysis in category_analysis:
                story.append(Paragraph(analysis, normal_style))
        
        story.append(Spacer(1, 20))
        
        # 财务分析
        story.append(Paragraph('综合财务分析与建议', heading_style))
        
        # 分析内容
        analysis_points = []
        if net_profit > 0:
            analysis_points.append('• 报告期内实现盈利，经营状况良好。')
        else:
            analysis_points.append('• 报告期内出现亏损，需要关注成本控制和收入增长。')
        
        if total_income > 0:
            profit_margin = net_profit / total_income * 100
            if profit_margin > 20:
                analysis_points.append('• 利润率超过20%，盈利能力优秀。')
            elif profit_margin > 10:
                analysis_points.append('• 利润率在10%-20%之间，盈利能力良好。')
            elif profit_margin > 0:
                analysis_points.append('• 利润率较低，建议优化成本结构。')
            else:
                analysis_points.append('• 出现亏损，建议重新评估业务模式。')
        
        # 月度趋势分析
        if len(month_data) >= 2:
            months = sorted(month_data.keys())
            recent_months = months[-3:] if len(months) >= 3 else months
            recent_profits = [month_data[m]['income'] - month_data[m]['expense'] for m in recent_months]
            
            if len(recent_profits) >= 2:
                if recent_profits[-1] > recent_profits[0]:
                    analysis_points.append('• 近期利润呈上升趋势，经营改善明显。')
                elif recent_profits[-1] < recent_profits[0]:
                    analysis_points.append('• 近期利润呈下降趋势，需要关注经营风险。')
                else:
                    analysis_points.append('• 近期利润保持稳定。')
        
        for point in analysis_points:
            story.append(Paragraph(point, normal_style))
        
        story.append(Spacer(1, 20))
        
        # 建议
        story.append(Paragraph('经营建议', heading_style))
        recommendations = [
            '• 定期监控关键财务指标，及时发现经营问题。',
            '• 加强成本控制，提高运营效率。',
            '• 多元化收入来源，降低经营风险。',
            '• 建立完善的财务预算和控制体系。',
            '• 关注现金流管理，确保资金链安全。'
        ]
        
        for rec in recommendations:
            story.append(Paragraph(rec, normal_style))
        
        # 生成PDF
        doc.build(story)
        output.seek(0)
        
        # 创建响应
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['X-Content-Type-Options'] = 'nosniff'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        
        # 生成文件名
        filename = f'financial_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        filename_utf8 = f'财务分析报告_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'\'{filename_utf8.encode("utf-8").decode("latin1")}"'
        
        return response
        
    except Exception as e:
        flash(f'PDF报告生成失败: {str(e)}', 'error')
        return redirect(url_for('statistics'))

@app.route('/receivables')
@login_required
def receivables():
    # 显示所有应收款记录，而不是只显示当前用户的
    receivables_list = Receivable.query.order_by(Receivable.created_at.desc()).all()
    
    # 计算统计数据
    total_amount = sum(r.amount for r in receivables_list)
    received_amount = sum(r.received_amount for r in receivables_list)
    pending_amount = total_amount - received_amount
    
    # 计算逾期金额
    overdue_amount = sum(r.remaining_amount for r in receivables_list if r.overdue_days > 0)
    
    total_count = len(receivables_list)
    
    return render_template('receivables.html', 
                         receivables=receivables_list,
                         total_amount=total_amount,
                         received_amount=received_amount,
                         pending_amount=pending_amount,
                         overdue_amount=overdue_amount,
                         total_count=total_count)

@app.route('/receivables/add', methods=['GET', 'POST'])
@edit_required
def add_receivable():
    if request.method == 'POST':
        title = request.form['title']
        amount = float(request.form['amount'])
        invoice_date_str = request.form.get('invoice_date')
        due_date_str = request.form.get('due_date')
        payment_terms = int(request.form.get('payment_terms', 30))
        contact_person = request.form.get('contact_person', '')
        contact_phone = request.form.get('contact_phone', '')
        contact_address = request.form.get('contact_address', '')
        notes = request.form.get('notes', '')
        
        # 生成应收款单号
        today = datetime.now()
        receivable_number = f"AR{today.strftime('%Y%m%d')}{str(Receivable.query.count() + 1).zfill(3)}"
        
        # 处理日期
        invoice_date = None
        due_date = None
        
        if invoice_date_str:
            try:
                invoice_date = datetime.strptime(invoice_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        if due_date_str:
            try:
                due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()
            except ValueError:
                # 如果没有指定到期日期但有开票日期，则根据账期计算
                if invoice_date:
                    due_date = invoice_date + timedelta(days=payment_terms)
        elif invoice_date:
            # 如果只有开票日期，根据账期计算到期日期
            due_date = invoice_date + timedelta(days=payment_terms)
        
        receivable = Receivable(
            receivable_number=receivable_number,
            title=title,
            amount=amount,
            invoice_date=invoice_date,
            due_date=due_date,
            payment_terms=payment_terms,
            contact_person=contact_person,
            contact_phone=contact_phone,
            contact_address=contact_address,
            notes=notes,
            user_id=current_user.id
        )
        
        db.session.add(receivable)
        db.session.commit()
        flash('应收款添加成功', 'success')
        return redirect(url_for('receivables'))
    
    return render_template('add_receivable.html')

@app.route('/receivables/<int:id>/receive', methods=['POST'])
@edit_required
def receive_receivable(id):
    receivable = Receivable.query.get_or_404(id)
    if receivable:
        received_amount = float(request.form.get('received_amount', 0))
        if received_amount > 0:
            receivable.received_amount += received_amount
            
            # 更新状态
            if receivable.received_amount >= receivable.amount:
                receivable.status = 'received'
                receivable.received_at = datetime.utcnow()
            else:
                receivable.status = 'partial'
            
            db.session.commit()
            flash(f'收款记录已更新，收款金额：¥{received_amount:.2f}', 'success')
        else:
            flash('收款金额必须大于0', 'error')
    else:
        flash('应收款不存在', 'error')
    
    return redirect(url_for('receivables'))

@app.route('/receivables/<int:id>/mark_received')
@edit_required
def mark_received(id):
    """快速标记为已收款"""
    receivable = Receivable.query.get_or_404(id)
    if receivable:
        receivable.received_amount = receivable.amount
        receivable.status = 'received'
        receivable.received_at = datetime.utcnow()
        db.session.commit()
        flash('应收款已标记为已收', 'success')
    else:
        flash('应收款不存在', 'error')
    
    return redirect(url_for('receivables'))

@app.route('/receivables/<int:id>/delete')
@edit_required
def delete_receivable(id):
    receivable = Receivable.query.get_or_404(id)
    if receivable:
        db.session.delete(receivable)
        db.session.commit()
        flash('应收款已删除', 'success')
    else:
        flash('应收款不存在', 'error')
    
    return redirect(url_for('receivables'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        
        if user and check_password_hash(user.password_hash, password) and user.is_active:
            login_user(user)
            return redirect(url_for('dashboard'))
        else:
            flash('用户名或密码错误', 'error')
    
    return render_template('login.html')

@app.route('/settings')
@admin_required
def settings():
    return render_template('settings.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# 仅在直接运行app.py时启动开发服务器（用于开发调试）
if __name__ == '__main__':
    with app.app_context():
        try:
            db.create_all()
            print("数据库表创建成功")

            if not User.query.filter_by(username='admin').first():
                admin = User(
                    username='admin',
                    email='admin@company.com',
                    password_hash=generate_password_hash('admin123'),
                    role='admin'
                )
                db.session.add(admin)
                db.session.commit()
                print("默认管理员账户已创建 - 用户名: admin, 密码: admin123")
            else:
                print("管理员账户已存在")
                
        except Exception as e:
            print(f"数据库连接或初始化失败: {e}")
            print("请检查MySQL连接配置和数据库是否存在")
            exit(1)
    
    print("\n注意: 当前使用Flask开发服务器")
    print("生产环境请使用: python wsgi.py 或 start_wsgi.bat")
    print("="*50)
    app.run(host='0.0.0.0', port=5085, debug=True)